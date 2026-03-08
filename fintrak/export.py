from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter


# ── Styles ────────────────────────────────────────────────────────────────────

_THIN = Side(style="thin")
_BORDER = Border(bottom=_THIN)
_HEADER_FONT = Font(bold=True, size=14)
_SECTION_FONT = Font(bold=True, size=11, color="2F5496")
_SUBTOTAL_FONT = Font(bold=True, size=11)
_CURRENCY_FMT = '#,##0.00'
_GREEN_FONT = Font(color="227447")
_RED_FONT = Font(color="C0392B")
_NET_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")


def export_pnl(pnl: dict, breakdown: list[dict], month: str, output_path: Path) -> Path:
    """Export P&L report as an Excel spreadsheet.

    Args:
        pnl: Result from analysis.profit_loss()
        breakdown: Result from analysis.expense_breakdown()
        month: Month string (YYYY-MM)
        output_path: Where to save the .xlsx file

    Returns:
        The path to the saved file.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "P&L Report"

    # Column widths
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 18

    row = 1

    # ── Title ─────────────────────────────────────────────────────────────
    ws.merge_cells("A1:B1")
    cell = ws.cell(row=row, column=1, value="Profit & Loss Report")
    cell.font = _HEADER_FONT
    cell.alignment = Alignment(horizontal="center")
    row += 1

    ws.merge_cells("A2:B2")
    cell = ws.cell(row=row, column=1, value=f"Period: {month}")
    cell.font = Font(size=11, italic=True, color="666666")
    cell.alignment = Alignment(horizontal="center")
    row += 1

    ws.merge_cells("A3:B3")
    cell = ws.cell(row=row, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    cell.font = Font(size=9, color="999999")
    cell.alignment = Alignment(horizontal="center")
    row += 2

    # ── Income ────────────────────────────────────────────────────────────
    row = _write_section_header(ws, row, "INCOME")
    for item in pnl["income_items"]:
        ws.cell(row=row, column=1, value=f"  {item['name']}")
        c = ws.cell(row=row, column=2, value=item["amount"])
        c.number_format = _CURRENCY_FMT
        c.font = _GREEN_FONT
        row += 1
    row = _write_subtotal(ws, row, "Total Income", pnl["total_income"], _GREEN_FONT)
    row += 1

    # ── Recurring Expenses ────────────────────────────────────────────────
    row = _write_section_header(ws, row, "EXPENSES — Recurring")
    for item in pnl["recurring_expense_items"]:
        ws.cell(row=row, column=1, value=f"  {item['name']}")
        c = ws.cell(row=row, column=2, value=item["amount"])
        c.number_format = _CURRENCY_FMT
        c.font = _RED_FONT
        row += 1
    row = _write_subtotal(ws, row, "Total Recurring", pnl["total_recurring_expenses"], _RED_FONT)
    row += 1

    # ── Credit Card Expenses ──────────────────────────────────────────────
    row = _write_section_header(ws, row, "EXPENSES — Credit Cards")
    for item in pnl["card_expense_items"]:
        ws.cell(row=row, column=1, value=f"  {item['card']}")
        c = ws.cell(row=row, column=2, value=item["amount"])
        c.number_format = _CURRENCY_FMT
        c.font = _RED_FONT
        row += 1
    row = _write_subtotal(ws, row, "Total Credit Cards", pnl["total_card_expenses"], _RED_FONT)
    row += 1

    # ── Total Expenses ────────────────────────────────────────────────────
    row = _write_subtotal(ws, row, "TOTAL EXPENSES", pnl["total_expenses"], _RED_FONT)
    row += 1

    # ── Net P/L ───────────────────────────────────────────────────────────
    net = pnl["net"]
    net_font = Font(bold=True, size=12, color="227447" if net >= 0 else "C0392B")
    a = ws.cell(row=row, column=1, value="NET PROFIT / LOSS")
    a.font = Font(bold=True, size=12)
    a.fill = _NET_FILL
    b = ws.cell(row=row, column=2, value=net)
    b.number_format = _CURRENCY_FMT
    b.font = net_font
    b.fill = _NET_FILL
    b.alignment = Alignment(horizontal="right")
    row += 2

    # ── Expense Breakdown Table + Chart ───────────────────────────────────
    if breakdown:
        row = _write_section_header(ws, row, "EXPENSE BREAKDOWN")
        breakdown_start = row

        # Headers
        ws.cell(row=row, column=1, value="Category").font = Font(bold=True)
        c = ws.cell(row=row, column=2, value="Amount")
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="right")
        row += 1

        for item in breakdown:
            ws.cell(row=row, column=1, value=item["label"])
            c = ws.cell(row=row, column=2, value=item["amount"])
            c.number_format = _CURRENCY_FMT
            c.alignment = Alignment(horizontal="right")
            row += 1

        breakdown_end = row - 1

        # Bar chart
        chart = BarChart()
        chart.type = "bar"
        chart.title = "Expense Breakdown"
        chart.y_axis.title = None
        chart.x_axis.title = None
        chart.legend = None
        chart.style = 10
        chart.width = 22
        chart.height = max(10, min(len(breakdown) * 1.2, 20))

        data = Reference(ws, min_col=2, min_row=breakdown_start, max_row=breakdown_end)
        cats = Reference(ws, min_col=1, min_row=breakdown_start + 1, max_row=breakdown_end)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)

        # Color the bars
        series = chart.series[0]
        series.graphicalProperties.solidFill = "2F5496"

        ws.add_chart(chart, f"D{breakdown_start}")

    # ── Print settings ────────────────────────────────────────────────────
    ws.print_title_rows = "1:3"
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    wb.save(output_path)
    return output_path


def _write_section_header(ws, row, title):
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = _SECTION_FONT
    ws.cell(row=row, column=2).border = _BORDER
    cell.border = _BORDER
    return row + 1


def _write_subtotal(ws, row, label, amount, font):
    a = ws.cell(row=row, column=1, value=label)
    a.font = _SUBTOTAL_FONT
    a.border = Border(top=_THIN)
    b = ws.cell(row=row, column=2, value=amount)
    b.number_format = _CURRENCY_FMT
    b.font = Font(bold=True, color=font.color)
    b.alignment = Alignment(horizontal="right")
    b.border = Border(top=_THIN)
    return row + 1
